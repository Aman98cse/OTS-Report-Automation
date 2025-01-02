import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

# Set up logging
logging.basicConfig(
    filename='raw_data/logs.log',  # File to store logs
    level=logging.INFO,  # Log level
    format='%(asctime)s - %(levelname)s - %(message)s'
)

file_path = 'raw_data/PVVNL_OTS_NOV-24_REPORT.csv'
logging.info(f"Reading the file name: {file_path}")
df = pd.read_csv(file_path)

logging.info("Adding the 'PAID_SLAB' column based of 'TOTAL_PAID' column")
df['PAID_SLAB'] = df['TOTAL_PAID'].apply(lambda x: "Below 50K" if x < 50000 else "50K to 1Lac" if (x >= 50000 and x < 100000) else "above 1Lac")

logging.info("Create a Pivot Table using pandas")
pivot_table = pd.pivot_table(df,
                             values=['TOTAL_PAID'],
                             index=['ZONE', 'CIRCLE', 'DIVISION_NAME'],
                             columns='PAID_SLAB',
                             aggfunc={'ZONE': 'count', 'TOTAL_PAID': 'sum'},
                             fill_value=0)

logging.info("Flatten multi-level columns")
pivot_table.columns = [f'{col}_{agg}' for col, agg in pivot_table.columns]

logging.info("Add total columns for TOTAL_PAID")
pivot_table['Total Sum of TOTAL_PAID'] = pivot_table[[
    'TOTAL_PAID_50K to 1Lac',
    'TOTAL_PAID_Below 50K',
    'TOTAL_PAID_above 1Lac'
]].sum(axis=1)

logging.info("Add total columns for ZONE")
pivot_table['Total Count of ZONE'] = pivot_table[[
    'ZONE_50K to 1Lac',
    'ZONE_Below 50K',
    'ZONE_above 1Lac'
]].sum(axis=1)

logging.info("Adding Subtotals")
# Reset index for easier manipulation
pivot_table = pivot_table.reset_index()

# Add subtotals for each CIRCLE
grouped = pivot_table.groupby(['ZONE', 'CIRCLE'])
subtotal_rows_circle = []

for (zone, circle), group in grouped:
    subtotal = group.sum(numeric_only=True)
    subtotal['ZONE'] = zone
    subtotal['CIRCLE'] = circle + " Total"
    subtotal['DIVISION_NAME'] = None
    subtotal_rows_circle.append(subtotal)

# Add subtotals for each ZONE
grouped_zone = pivot_table.groupby(['ZONE'], sort=False)
subtotal_rows_zone = []

for zone, group in grouped_zone:
    zone_name = zone if isinstance(zone, str) else zone[0]  # Handle tuple if it appears
    subtotal = group.sum(numeric_only=True)
    subtotal['ZONE'] = f"{zone_name} Total"
    subtotal['CIRCLE'] = None
    subtotal['DIVISION_NAME'] = None
    subtotal_rows_zone.append(subtotal)

pivot_table = pd.concat([pivot_table, pd.DataFrame(subtotal_rows_zone)], ignore_index=True)
pivot_table = pd.concat([pivot_table, pd.DataFrame(subtotal_rows_circle)], ignore_index=True)

# Sort the table to ensure totals appear correctly
pivot_table = pivot_table.sort_values(by=['ZONE', 'CIRCLE', 'DIVISION_NAME'], key=lambda x: x.fillna(''))

# Avoid repeating ZONE and CIRCLE values for the same group
pivot_table[['ZONE', 'CIRCLE']] = pivot_table[['ZONE', 'CIRCLE']].where(pivot_table[['ZONE', 'CIRCLE']].ne(pivot_table[['ZONE', 'CIRCLE']].shift()))

logging.info("Saving the resultant file into an excel sheet")
pivot_file_path = 'raw_data/OTS_TURN_UP_CATEGORISATION_with_totals.xlsx'
pivot_table.to_excel(pivot_file_path, sheet_name='Pivot Table', index=False)

# Autofit columns
logging.info("Autofitting columns in the Excel sheet")
wb = load_workbook(pivot_file_path)
sheet = wb.active

for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column letter (e.g., A, B, C)
    for cell in col:
        try:
            # Calculate the maximum content length
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2  # Add some padding
    sheet.column_dimensions[column].width = adjusted_width


# Add borders to the table
logging.info("Adding borders to the table")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal='center', vertical='center')

for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_alignment


# Save the updated Excel file
wb.save(pivot_file_path)

logging.info(f"Successfully created the {pivot_file_path}")
print(f"Successfully created the {pivot_file_path}")
